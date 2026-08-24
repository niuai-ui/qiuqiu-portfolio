from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "作品信息.xlsx"
DEFAULT_LOG_ROOT = Path(r"E:\自媒体 湫湫sims日志\功能模组")
IGNORED_DIRECTORIES = {"待发布", "【飞书作品集】", "其他功能模组"}


def key(author: str, english_name: str) -> str:
    value = f"{author}-{english_name}"
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-zA-Z0-9]+", "-", normalized).strip("-").lower()


def parse_folder_name(name: str) -> tuple[str, str]:
    parts = [part.strip() for part in name.split(" - ")]
    if parts and re.fullmatch(r"\d{4}", parts[-1]):
        parts.pop()
    if len(parts) >= 3:
        return parts[0], parts[1]
    if len(parts) == 2 and "_" in parts[0]:
        author, english_name = parts[0].split("_", 1)
        return author.strip(), english_name.strip()
    raise ValueError(f"无法识别日志文件夹名称：{name}")


def source_folders(log_root: Path) -> list[Path]:
    if not log_root.is_dir():
        raise FileNotFoundError(f"找不到日志源目录：{log_root}")
    direct = [
        path for path in log_root.iterdir()
        if path.is_dir() and path.name not in IGNORED_DIRECTORIES
    ]
    other_root = log_root / "其他功能模组"
    nested = [path for path in other_root.iterdir() if path.is_dir()] if other_root.is_dir() else []
    return sorted(direct + nested, key=lambda path: path.name.casefold())


def workbook_records() -> dict[str, str]:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    sheet = workbook["作品信息"]
    headers = [cell.value for cell in next(sheet.iter_rows())]
    author_index = headers.index("原作者名字")
    english_index = headers.index("模组英文名")
    records: dict[str, str] = {}
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None for value in row):
            continue
        author = str(row[author_index] or "").strip()
        english_name = str(row[english_index] or "").strip()
        record_key = key(author, english_name)
        if not record_key:
            raise ValueError(f"Excel 第 {row_number} 行缺少作者或英文模组名")
        if record_key in records:
            raise ValueError(f"Excel 存在重复记录：第 {row_number} 行 {author} - {english_name}")
        records[record_key] = f"{author} - {english_name}"
    return records


def check_sync(log_root: Path = DEFAULT_LOG_ROOT) -> None:
    source: dict[str, str] = {}
    for folder in source_folders(log_root):
        author, english_name = parse_folder_name(folder.name)
        record_key = key(author, english_name)
        if record_key in source:
            raise ValueError(f"日志源存在重复模组：{folder.name}")
        source[record_key] = folder.name

    workbook = workbook_records()
    missing_in_workbook = [source[item] for item in sorted(source.keys() - workbook.keys())]
    missing_in_logs = [workbook[item] for item in sorted(workbook.keys() - source.keys())]
    if missing_in_workbook or missing_in_logs:
        lines = ["日志目录与作品信息.xlsx 不一致："]
        if missing_in_workbook:
            lines.append("  日志中有、Excel 中缺少：" + "；".join(missing_in_workbook))
        if missing_in_logs:
            lines.append("  Excel 中有、日志中缺少：" + "；".join(missing_in_logs))
        raise ValueError("\n".join(lines))
    print(f"同步检查通过：日志源与 Excel 均为 {len(source)} 个模组")


def main() -> None:
    parser = argparse.ArgumentParser(description="比较日志源目录与作品信息.xlsx，防止漏同步或重复记录。")
    parser.add_argument("--log-root", type=Path, default=DEFAULT_LOG_ROOT)
    args = parser.parse_args()
    check_sync(args.log_root)


if __name__ == "__main__":
    main()
