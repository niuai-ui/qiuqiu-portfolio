from __future__ import annotations

import json
import re
import shutil
import unicodedata
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
CONTENT = ROOT / "content"
DIST = ROOT / "dist"
SOURCE = ROOT / "site"
WORKBOOK = ROOT / "作品信息.xlsx"
REQUIRED = [
    "状态", "模组英文名", "模组中文名", "原作者名字", "原作者网址链接",
    "汉化发布日期", "汉化更新日期", "前置说明", "放置说明", "封面路径",
    "百度网盘链接整体", "类别",
]
ALLOWED_STATUSES = {"已发布", "草稿", "下架"}
ALLOWED_CATEGORIES = {"人物特征", "用地特征", "职业", "覆盖替换", "游戏玩法", "其他"}
ALLOWED_PLACEMENTS = {"必须放第一层", "无需放第一层"}
SPECIAL_IMAGES = {"favicon.png", "og.jpg"}


def text(value):
    return "" if value is None else str(value).strip()


def iso_date(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    value = text(value)
    return value[:10] if value else ""


def require_date(value, row_number, field_name):
    if not isinstance(value, (datetime, date)):
        raise ValueError(f"第 {row_number} 行的“{field_name}”必须是真正的 Excel 日期，不能是文本")
    return value.date() if isinstance(value, datetime) else value


def require_http_url(value, row_number, field_name):
    url = text(value)
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError(f"第 {row_number} 行的“{field_name}”必须是有效的 HTTP(S) 网址")
    return url


def slug(value):
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^a-zA-Z0-9]+", "-", normalized).strip("-").lower()
    return normalized or "work"


def split_download(value):
    """把“链接: URL 提取码: CODE”拆成网址和提取码，便于网页分别渲染。"""
    raw = text(value)
    if not raw:
        return "", ""
    url_match = re.search(r"https?://[^\s，,；;]+", raw)
    url = url_match.group(0).rstrip("。.、") if url_match else ""
    code_match = re.search(r"(?:提取码|密码|pwd)\s*[:：]?\s*([0-9A-Za-z]{4})", raw)
    code = code_match.group(1) if code_match else ""
    return url, code


def load_works():
    workbook = load_workbook(WORKBOOK, data_only=False)
    if "作品信息" not in workbook.sheetnames:
        raise ValueError("Excel 中缺少“作品信息”工作表")
    sheet = workbook["作品信息"]
    headers = [text(cell.value) for cell in sheet[1]]
    if headers != REQUIRED:
        raise ValueError(
            "Excel 必须严格使用 12 个固定字段并保持规定顺序。\n"
            f"应为：{'、'.join(REQUIRED)}\n"
            f"实际：{'、'.join(headers)}"
        )
    positions = {name: index for index, name in enumerate(headers)}
    works, ids = [], set()
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None for value in cells):
            continue
        item = {name: cells[index] if index < len(cells) else None for name, index in positions.items()}
        status = text(item.get("状态"))
        if status not in ALLOWED_STATUSES:
            raise ValueError(f"第 {row_number} 行的状态无效：{status or '空白'}")
        if status != "已发布":
            continue
        required_fields = [
            "模组英文名", "模组中文名", "原作者名字", "原作者网址链接",
            "汉化发布日期", "汉化更新日期", "前置说明", "放置说明", "封面路径", "类别",
        ]
        missing_values = [name for name in required_fields if not text(item.get(name))]
        if missing_values:
            raise ValueError(f"第 {row_number} 行缺少已发布作品必填内容：{'、'.join(missing_values)}")
        english = text(item.get("模组英文名"))
        author_name = text(item.get("原作者名字"))
        original_url = require_http_url(item.get("原作者网址链接"), row_number, "原作者网址链接")
        release_date = require_date(item.get("汉化发布日期"), row_number, "汉化发布日期")
        updated_date = require_date(item.get("汉化更新日期"), row_number, "汉化更新日期")
        if updated_date < release_date:
            raise ValueError(f"第 {row_number} 行的汉化更新日期不能早于汉化发布日期")
        placement = text(item.get("放置说明"))
        if placement not in ALLOWED_PLACEMENTS:
            raise ValueError(f"第 {row_number} 行的放置说明无效：{placement}")
        category = text(item.get("类别"))
        if category not in ALLOWED_CATEGORIES:
            raise ValueError(f"第 {row_number} 行的类别无效：{category}")
        work_id = slug(f"{author_name}-{english}")
        if work_id in ids:
            raise ValueError(f"第 {row_number} 行生成的作品标识重复：{work_id}")
        ids.add(work_id)
        download_url, download_code = split_download(item.get("百度网盘链接整体"))
        if text(item.get("百度网盘链接整体")) and not download_url:
            raise ValueError(f"第 {row_number} 行的百度网盘单元格里找不到有效网址")
        cover = text(item.get("封面路径"))
        cover_path = Path(cover)
        if (
            cover_path.is_absolute()
            or ".." in cover_path.parts
            or len(cover_path.parts) != 2
            or cover_path.parts[0] != "images"
            or cover_path.name in SPECIAL_IMAGES
        ):
            raise ValueError(f"第 {row_number} 行的封面路径必须是 images/文件名，且不能使用站点图标：{cover}")
        cover_file = CONTENT / cover_path
        if not cover_file.is_file():
            raise FileNotFoundError(f"第 {row_number} 行封面不存在：content/{cover}")
        with Image.open(cover_file) as image:
            if image.width * 4 != image.height * 3:
                raise ValueError(f"第 {row_number} 行封面不是标准 3:4：{cover}（{image.width}×{image.height}）")
        works.append({
            "id": work_id,
            "title": text(item.get("模组中文名")),
            "englishTitle": english,
            "author": author_name or "未知作者",
            "originalUrl": original_url,
            "date": iso_date(release_date),
            "updated": iso_date(updated_date),
            "dependency": text(item.get("前置说明")),
            "placement": placement,
            "localization": "繁简汉化",
            "sourceImage": cover,
            "download": download_url,
            "downloadCode": download_code,
            "category": category,
        })
    works.sort(key=lambda work: (work["date"], work["title"]), reverse=True)
    return works


def build_responsive_covers(works):
    output_dir = DIST / "images" / "covers"
    output_dir.mkdir(parents=True, exist_ok=True)
    original_bytes = 0
    optimized_bytes = 0
    for work in works:
        source = CONTENT / work.pop("sourceImage")
        original_bytes += source.stat().st_size
        with Image.open(source) as opened:
            image = opened.convert("RGB")
            generated = {}
            for width in (480, 960):
                resized = image.resize((width, width * 4 // 3), Image.Resampling.LANCZOS)
                filename = f"{work['id']}-{width}.webp"
                target = output_dir / filename
                resized.save(target, "WEBP", quality=82, method=6, optimize=True)
                optimized_bytes += target.stat().st_size
                generated[width] = f"images/covers/{filename}"
        work["imageSmall"] = generated[480]
        work["imageLarge"] = generated[960]
        work["image"] = generated[960]
    return original_bytes, optimized_bytes


def build():
    works = load_works()
    expected_dist = ROOT / "dist"
    if DIST != expected_dist or DIST.parent != ROOT:
        raise RuntimeError(f"拒绝清理异常输出目录：{DIST}")
    if DIST.exists():
        shutil.rmtree(DIST)
    shutil.copytree(SOURCE, DIST)
    (DIST / "images").mkdir(exist_ok=True)
    for filename in ("favicon.png", "og.jpg"):
        shutil.copy2(CONTENT / "images" / filename, DIST / "images" / filename)
    original_bytes, optimized_bytes = build_responsive_covers(works)
    (DIST / "data.json").write_text(json.dumps(works, ensure_ascii=False, indent=2), encoding="utf-8")
    (DIST / ".nojekyll").write_text("", encoding="utf-8")
    print(f"网站已生成：{len(works)} 个已发布作品 → {DIST}")
    print(f"封面体积：{original_bytes / 1024 / 1024:.2f} MB → {optimized_bytes / 1024 / 1024:.2f} MB（480/960 WebP 合计）")


if __name__ == "__main__":
    build()
